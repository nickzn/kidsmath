#!/usr/bin/env python3


import argparse
import random
import math
import ast
import operator as op
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side


def main():
    parser = argparse.ArgumentParser(description='Generate kids math tests')
    parser.add_argument('--debug', '-d', dest='debug',
                        action='store_true', help='debug mode')
    args = parser.parse_args()
    lower_limit = 1
    upper_limit = 20
    operators = ['+', '-', '*', '/']
    # operators = ['+', '-']
    n_numbers = 4
    n_tests = 100
    filename = 'kidsmath.xlsx'
    split_num = int(7 / n_numbers)
    if split_num == 0:
        split_num = 1

    tests, results = gen_test(operators,
                              upper_limit, lower_limit, n_numbers, n_tests)
    gen_xlsx(filename, tests, results, n_numbers, split_num)
    print('%s generated!\n' % filename)
    if args.debug:
        for i, test in enumerate(tests):
            try:
                answer = eval_expr(test)
            except SyntaxError:
                answer = 'f'
            if answer == results[i]:
                f = 'correct'
            else:
                f = 'incorrect'
            print('%-50s%-10s' % ('%s = %s' % (test, results[i]), f))
    return None


def gen_xlsx(filename, tests, results, n_numbers, split_num):
    wb = Workbook()
    ws = wb.active
    data = []
    row = []
    ft = Font(size=16)
    bd = Border(bottom=Side(border_style='thin'))
    n_col_every_formula = 5  # [formula, =, '', answer, '']
    max_columns = split_num * n_col_every_formula
    formula_columns = list(range(1, max_columns + 1, n_col_every_formula))

    for i, test in enumerate(tests):
        test = re.sub(r'\*', '×', test)
        test = re.sub(r'\/', '÷', test)
        row.extend([test, '=', '', results[i], ''])
        if (i + 1) % split_num == 0:
            data.append(row)
            row = []
    for i, d in enumerate(data):
        ws.append(d)
        for j in formula_columns:
            # j + 2 is result column
            for k in (j, j + 2):
                c = ws.cell(row=i + 1, column=k)
                c.font = ft
                c.border = bd
    adjust_column_width(ws, formula_columns, split_num)
    wb.save(filename=filename)


def adjust_column_width(ws, formula_columns, split_num):
    equal_sign_columns = [i + 1 for i in formula_columns]
    hide_columns = [i + 3 for i in formula_columns]
    seperator_columns = [i + 4 for i in formula_columns]
    for i in hide_columns:
        # hide answer column and set width of result column
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions.group(letter, letter, hidden=True)
        letter = ws.cell(row=1, column=i - 1).column_letter
        ws.column_dimensions[letter].width = calculate_width(
            tuple(ws.columns)[i - 1]) * 2

    for i in equal_sign_columns:
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = 3
    for i in seperator_columns:
        letter = ws.cell(row=1, column=i).column_letter
        # adjust width of seperator
        ws.column_dimensions[letter].width = 24 / split_num
    for i in formula_columns:
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = calculate_width(
            tuple(ws.columns)[i - 1])

    for i in range(1, len(tuple(ws.rows)) + 1):
        ws.row_dimensions[i].height = 37


def calculate_width(cells):
    return max(len(str(cell.value)) for cell in cells) + 2


def gen_test(operators, upper_limit, lower_limit, n_numbers, n_tests):
    func_of_operator = {
        '+': numbers_for_plus,
        '-': numbers_for_minus,
        '*': numbers_for_multiple,
        '/': numbers_for_divide,
    }
    all_tests = []
    all_results = []
    i = 0
    while i < n_tests:
        numbers, generated_operators, result = gen_random(
            operators, func_of_operator, upper_limit, lower_limit, n_numbers
        )
        all_tests.append(gen_formula(numbers, generated_operators))
        all_results.append(result)
        i += 1
    return all_tests, all_results


def gen_formula(numbers, operators):
    if len(operators) == 1:
        operator = ' %s ' % operators[0]
        return operator.join([str(i) for i in numbers])
    formula = ''
    max_n = len(operators) - 1
    n = max_n
    while n >= 0:
        parenthesis_1 = ''
        parenthesis_2 = ''
        prev_operator = safe_list_get(operators, n - 1)
        if (prev_operator in ('-', '/') or
            (prev_operator in ('*', '/') and
             operators[n] in ('+', '-'))):
            parenthesis_1 = '('
            parenthesis_2 = ')'
        if n == max_n:
            num_2nd = numbers[n + 1]
        elif n == 0:
            num_2nd = formula
            parenthesis_1 = ''
            parenthesis_2 = ''
        else:
            num_2nd = formula
        formula = '%s%s %s %s%s' % (
            parenthesis_1, numbers[n], operators[n], num_2nd, parenthesis_2
        )
        n -= 1
    return formula


def safe_list_get(l, idx, default=None):
    try:
        return l[idx]
    except IndexError:
        return default
    return default


def gen_random(
    operators, func_of_operator,
    upper_limit, lower_limit, n_numbers
):
    numbers = []
    generated_operators = []
    target = random.randint(lower_limit, upper_limit)
    i = 0
    remain = target
    while i < n_numbers - 1:
        operator = random.choice(operators)
        if operator == '/' and remain == 0:
            operator = '*'
        n, remain = func_of_operator[operator](remain,
                                               lower_limit, upper_limit)
        numbers.append(n)
        generated_operators.append(operator)
        i += 1
    numbers.append(remain)
    return numbers, generated_operators, target


def numbers_for_plus(target, lower_limit, upper_limit):
    if target:
        n1 = random.randint(1, target)
        if n1 == target:
            n1 -= 1
    else:
        n1 = 0
    n2 = target - n1
    return n1, n2


def numbers_for_minus(target, lower_limit, upper_limit):
    n1 = random.randint(target, upper_limit)
    n2 = n1 - target
    return n1, n2


def numbers_for_multiple(target, lower_limit, upper_limit):
    n1 = int(random.choice(divisors(target)))
    n2 = int(target / n1)
    return n1, n2


def numbers_for_divide(target, lower_limit, upper_limit):
    multiple = int(upper_limit / target)
    rand = random.choice(range(1, multiple + 1))
    n1 = target * rand
    n2 = rand
    return n1, n2


def divisors(n):
    divs = [1]
    for i in range(2, int(math.sqrt(n)) + 1):
        if n % i == 0:
            divs.extend([i, n/i])
    if n:
        divs.extend([n])
    return list(set(divs))


def eval_expr(expr):
    return eval_(ast.parse(expr, mode='eval').body)


def eval_(node):
    # supported operators
    operators = {ast.Add: op.add, ast.Sub: op.sub, ast.Mult: op.mul,
                 ast.Div: op.truediv, ast.Pow: op.pow, ast.BitXor: op.xor,
                 ast.USub: op.neg}

    if isinstance(node, ast.Num):  # <number>
        return node.n
    elif isinstance(node, ast.BinOp):  # <left> <operator> <right>
        return operators[type(node.op)](eval_(node.left), eval_(node.right))
    elif isinstance(node, ast.UnaryOp):  # <operator> <operand> e.g., -1
        return operators[type(node.op)](eval_(node.operand))
    else:
        raise TypeError(node)


if __name__ == '__main__':
    main()
